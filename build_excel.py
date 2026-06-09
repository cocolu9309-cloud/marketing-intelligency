from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
import copy

# Load template
wb = load_workbook(r'D:\工作\网翔\课题方案\用户运营\Remarketing\20260526_Callie郑锶婷_美国站2026年5月赠品活动ins新媒体素材设计需求表.xlsx')
ws = wb.active

# Helper function for long text
def set_cell(ws, cell_ref, value, wrap=True):
    cell = ws[cell_ref]
    cell.value = value
    cell.alignment = Alignment(wrap_text=wrap, vertical='top')

# Row 2: Activity info
ws['B2'] = '国际爱狗日Remarketing广告素材需求表'
ws['D2'] = '用户运营产品活动标题----'

# Row 3: Date info
ws['B3'] = '业务负责人：'
ws['D3'] = '2026/08/26'
ws['F3'] = '上新时间：'
ws['H3'] = '2026.08.26'

# Row 4: Content info
ws['B4'] = '业务定位：'
ws['D4'] = '内容丰富度强·4个AD创意方向·产品矩阵'
ws['F4'] = '规格/产品特点：'
ws['H4'] = '25个产品SKU·4套视频脚本'

# Row 5: Objective
ws['B5'] = '业务目标：'
ws['D5'] = '通过多素材Meta Ads+Google PMax投放，测试不同创意方向，找出数据表现最好的方向并倾斜预算，ROAS≥1.45'
ws['F5'] = '结论/鬼畜/平泽：'
ws['H5'] = '主要为独立站市场'

# Row 6: Target audience
ws['B6'] = '目标族群：'
ws['D6'] = 'Callie.com老客户（已购买宠物定制产品）'
ws['F6'] = '使用平台：'
ws['H6'] = 'Meta Ads / Google PMax'

# AD1 content in Row 9
ad1_content = """【视频脚本】时长：15-21秒 | 画幅：9:16（1080×1920）| 格式：MP4

【HOOK】0:00-0:02
画面：大字动画 "AUG 26 + 国际爱狗日图标" + 爪印
声音：紧迫音乐前奏，低频心跳
目的：吸引注意力，唤起节日感

【视频主干】0:02-0:15
0:02-0:04 宠物眼部特写，慢动作，字幕："She never says it. But she shows it every day."
0:04-0:07 拉远：宠物依偎在主人怀中，温馨场景，字幕："International Dog Day"
0:07-0:11 产品展示：宠物照片印刷在毛毯/画布上，镜头展示产品，字幕："Turn your best friend into a treasure"
0:11-0:15 主人用手机扫描产品，落地页加载，字幕："Swipe up to shop"

【CTA】0:15-0:21
0:15-0:18 产品图 + "Starting at $18" + 按钮："Honor the one who loves you →"
0:18-0:21 品牌水印，文字："Free Shipping on $50+ | Only for Callie Pet Lovers"

【视觉风格】
暖色调（米色、浅棕、奶油白），柔焦背景，宠物毛发细节清晰

【字幕要求】
所有文字内嵌字幕，字体≥60pt（主标题）/≥48pt（价格）/≥36pt（CTA）

【产品画面要求】
宠物照片+产品实物结合，宠物照片是产品印刷内容（不是单独宠物照）

【音乐建议】
温情钢琴/轻柔吉他，情感共鸣感"""

set_cell(ws, 'E9', ad1_content)

# AD1 SKUs
ws['F9'] = """AD1产品（10个SKU）
CAJJ261101 $27 宠物肖像·画布印刷
CAJJ252584 $23 宠物爪印·毛毯
CAJJ252399 $25 宠物水彩肖像·毛毯
CAJJ252410 $23 宠物搞笑面孔·毛毯
CAJJ252078 $35 4宠照片·毛毯
CAJJ250722 $18 5D钻石画DIY套装
CAJJ251853 $23 多色爱心·毛毯
PL251744 $23 3D卡通狗·折叠收纳篮
CAJJ251357 $23 多色宠物肖像·毛毯
CAJJ251316 $20 宠物爪印·抱枕套"""

ws['I9'] = '15-21秒'

# AD2 content in Row 13 (skip some rows for spacing)
ad2_row = 13
ws.cell(row=ad2_row, column=1).value = '视频素材创意'
ws.cell(row=ad2_row, column=2).value = 2

ad2_content = """【视频脚本】时长：15-21秒 | 画幅：9:16（1080×1920）| 格式：MP4

【HOOK】0:00-0:02
画面：日历翻到8月26日，圈出日期 + "Aug 26. International Dog Day." 大字
声音：轻快音乐前奏
目的：唤起国际爱狗日节日感

【视频主干】0:02-0:17
0:02-0:05 主人拿着产品（宠物睡衣/领带/帽子）对着镜子自拍，字幕："Dog mom energy. 365 days a year."
0:05-0:09 场景快切：遛狗 / 宠物窝在印有自己照片的毯子上 / 地垫放在门口，字幕："This is who I am."
0:09-0:13 产品使用展示快切：睡衣穿上身、领带系好、帽子戴上（3个产品，2-3秒切换）
0:13-0:17 场景延伸：宠物戴着配件跑，主人自信走姿

【CTA】0:17-0:21
0:17-0:20 多产品展示 + "Starting at $20" + 按钮："Show Your Dog Mom Pride →"
0:20-0:21 品牌水印，文字："Free Shipping on $50+ | Aug 26 Exclusive"

【视觉风格】
活力色调（珊瑚粉、薄荷绿、浅蓝），自然光拍摄，生活场景（非影棚），主人和宠物自然互动

【字幕要求】
所有文字内嵌字幕，字体≥60pt（主标题）/≥48pt（价格）/≥36pt（CTA）

【产品画面要求】
产品必须上身/使用中（睡衣穿在身上、领带系好、帽子戴头上），不是产品静物图

【音乐建议】
轻快流行/活力节拍，年轻养宠人风格"""

set_cell(ws, f'E{ad2_row}', ad2_content)

ws[f'F{ad2_row}'] = """AD2产品（8个SKU）
PL260891 $23 宠物照片领带
CAJJ260304 $33 宠物照片·花彩虹天使翼·花园灯
PG5308 $29 多色宠物肖像·睡衣裤
CAFS250460 $46 多色宠物肖像·长袖家居服套装
CAJJ251832 $26 1-3只3D卡通狗·防滑地垫
PG3270 $29 刺绣宠物照片·棒球帽
CAPS251274 $20 3D卡通狗·磁吸手机支架
CAFS250177 $33 搞笑狗猫宠·短袖女睡衣套装"""

ws[f'I{ad2_row}'] = '15-21秒'

# AD3 content
ad3_row = 17
ws.cell(row=ad3_row, column=1).value = '视频素材创意'
ws.cell(row=ad3_row, column=2).value = 3

ad3_content = """【视频脚本】时长：15-21秒 | 画幅：9:16（1080×1920）| 格式：MP4

【HOOK】0:00-0:02
画面：礼物盒特写，丝带上印有爪印图案，字幕："AUG 26"
声音：轻快礼物音效
目的：节日 + 礼物双重吸引力

【视频主干】0:02-0:17
0:02-0:05 开箱：宠物凑过来，好奇嗅闻，人物表情惊喜，字幕："International Dog Day. The perfect gift for the pet lover."
0:05-0:09 产品展示：马克杯（宠物照片）、宠物护照相册、窗挂装饰依次展示（每个2-3秒）
0:09-0:12 产品放在温馨场景中（桌上、窗边、墙上），字幕："Because they understand what 'dog person' really means."
0:12-0:17 礼物包装感展示：产品放入礼盒，丝带系好，突出"送礼仪式感"

【CTA】0:17-0:21
0:17-0:20 礼盒+产品图 + "$18-$29" + 按钮："Find the Perfect Gift →"
0:20-0:21 品牌水印，文字："Free Shipping on $50+ | Aug 26 Gift Guide"

【视觉风格】
礼物感强（礼盒、丝带、包装纸元素），暖金色调（礼物氛围），产品主体清晰，人物表情自然惊喜

【字幕要求】
所有文字内嵌字幕，字体≥60pt（主标题）/≥48pt（价格）/≥36pt（CTA）

【产品画面要求】
产品必须有礼物感（礼盒/丝带/包装元素），不是裸产品展示

【音乐建议】
轻快礼物感/钢琴曲，温馨收礼氛围"""

set_cell(ws, f'E{ad3_row}', ad3_content)

ws[f'F{ad3_row}'] = """AD3产品（5个SKU）
CAPS252666 $25 宠物护照（护照风格相册）
CAJJ251421 $29 猫狗轮廓·12oz陶瓷马克杯
CAJJ251727 $26 花宠物照片·陶瓷花盆
PG3411 $18 彩色玻璃风格·窗挂装饰
CAPS252548 $25 骨头/鸡腿/爪印·宠物名字标签"""

ws[f'I{ad3_row}'] = '15-21秒'

# AD4 content
ad4_row = 21
ws.cell(row=ad4_row, column=1).value = '视频素材创意'
ws.cell(row=ad4_row, column=2).value = 4

ad4_content = """【视频脚本】时长：15-21秒 | 画幅：9:16（1080×1920）| 格式：MP4

【策略说明】
用低于$20的低价产品吸引价格敏感的老客户点击进入落地页，通过落地页产品推荐完成升级转化（低价引流+产品升级）

【HOOK】0:00-0:02
画面：大字 "AUG 26 | International Dog Day" + 爪印图标
声音：电子音乐前奏，紧迫感
目的：唤起节日感，吸引注意力

【视频主干】0:02-0:19
0:02-0:05 文字动画："Great gifts start at just $11" + "For pet lovers. From Callie."
0:05-0:09 产品快闪展示（3个产品快速切换）：狗ID标签$11 / 窗挂装饰$18 / 牵狗绳$20，每个1-2秒
0:09-0:12 产品特写 + 价格大字，字幕："Tiny price. Tail-wagging joy."
0:12-0:16 场景：宠物戴着ID标签奔跑 / 主人用牵狗绳遛狗，字幕："Everything for your best friend"
0:16-0:19 落地页加载，字幕："Discover more at Callie"

【CTA】0:19-0:21
0:19-0:20 低价产品 + "From $11" + 按钮："Start Shopping →"
0:20-0:21 品牌水印，文字："Free Shipping on $50+ | Callie Pet Collection"

【视觉风格】
高对比（深色背景+白字/红字），大字价格（倒计时感），产品小但清晰，紧迫感

【字幕要求】
所有文字内嵌字幕，字体≥60pt（主标题）/≥48pt（价格）/≥36pt（CTA）

【产品画面要求】
产品放在生活场景中使用（宠物戴着跑、主人遛狗），不是静物图

【音乐建议】
紧迫电子/倒计时音效，节奏感强"""

set_cell(ws, f'E{ad4_row}', ad4_content)

ws[f'F{ad4_row}'] = """AD4产品（6个SKU）— 低价引流策略
PG2260 $11 闪光骨头·狗ID标签
PG3411 $18 彩色玻璃风格·窗挂装饰
CAJJ250722 $18 5D钻石画DIY套装
CAJJ251316 $20 宠物爪印·抱枕套
CAPS251274 $20 3D卡通狗·磁吸手机支架
CAPS252024 $20 多色·卡通宠物·伸缩牵狗绳

策略说明：低价产品吸引价格敏感老客户，点击落地页后通过产品推荐完成升级转化"""

ws[f'I{ad4_row}'] = '15-21秒'

# Save file
output_path = r'D:\工作\网翔\课题方案\用户运营\Remarketing\20260826_Callie国际爱狗日Remarketing广告素材设计需求表.xlsx'
wb.save(output_path)
print(f"File saved to: {output_path}")