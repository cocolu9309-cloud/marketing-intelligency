# build_excel.py
"""
生成两个 Excel 文件：
1. 内容生成包.xlsx — 英文文案 + 设计底层逻辑
2. 设计对接文档.xlsx — 完整设计参数 + 关键帧参考图
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
from typing import Dict, List
import logging

logger = logging.getLogger(__name__)


def scale_image(img: XLImage, max_width: int, max_height: int) -> XLImage:
    """Scale image while preserving aspect ratio."""
    w, h = img.width, img.height
    scale = min(max_width / w, max_height / h)
    img.width, img.height = int(w * scale), int(h * scale)
    return img


def build_content_pack(data: Dict, output_path: str):
    """
    生成内容生成包.xlsx
    Sheet1: 内容总览
    Sheet2: 英文文案 + 11条Hashtag
    Sheet3: 关键帧描述
    Sheet4: 产品匹配
    """
    wb = Workbook()

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    GOLD_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    SUBHEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(name="Arial", size=10)
    BOLD_FONT = Font(name="Arial", bold=True, size=10)
    thin = Side(style="thin", color="BFBFBF")
    THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_cell(cell, fill=None, font=None, alignment=None, border=None):
        if fill: cell.fill = fill
        if font: cell.font = font
        if alignment: cell.alignment = alignment
        if border: cell.border = border

    platform = data.get("platform", "Instagram")

    # Sheet1: 内容总览
    ws1 = wb.active
    ws1.title = "内容总览"
    ws1.merge_cells("A1:C1")
    c = ws1.cell(1, 1, f"Callie 社媒内容包 — {platform}")
    style_cell(c, fill=HEADER_FILL, font=Font(name="Arial", bold=True, color="FFFFFF", size=14),
               alignment=Alignment(horizontal="center", vertical="center"), border=THIN_BORDER)
    ws1.row_dimensions[1].height = 30
    ws1.merge_cells("A2:C2")
    p = ws1.cell(2, 1, f"产品：{data.get('product_name', '')} | 平台：{platform}")
    style_cell(p, fill=ALT_ROW_FILL, font=BODY_FONT,
               alignment=Alignment(horizontal="left", vertical="center", wrap_text=True), border=THIN_BORDER)
    ws1.row_dimensions[2].height = 25
    for i, h in enumerate(["板块", "内容", "备注"], 1):
        c = ws1.cell(3, i, h)
        style_cell(c, fill=HEADER_FILL, font=HEADER_FONT,
                   alignment=Alignment(horizontal="center", vertical="center"), border=THIN_BORDER)
    rows = [
        ("Big Idea", f"{data.get('big_idea_en', '')}\n{data.get('big_idea_cn', '')}", ""),
        ("品牌角度", f"{data.get('brand_angle_en', '')}\n{data.get('brand_angle_cn', '')}", ""),
        ("产品关联度", data.get("product_tie_in", ""), ""),
        ("英文贴文", data.get("caption", ""), "可直接发布"),
        ("Hashtag", " ".join(data.get("hashtags", [])) if isinstance(data.get("hashtags"), list) else data.get("hashtags", ""), ""),
        ("行动号召 (CTA)", data.get("cta", ""), ""),
        ("品牌安全检查", data.get("brand_safety_check", ""), ""),
    ]
    for i, (label, value, note) in enumerate(rows, 4):
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        c1 = ws1.cell(i, 1, label)
        style_cell(c1, fill=SUBHEADER_FILL, font=SUBHEADER_FONT,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True), border=THIN_BORDER)
        c2 = ws1.cell(i, 2, value)
        style_cell(c2, fill=fill, font=BODY_FONT,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True), border=THIN_BORDER)
        c3 = ws1.cell(i, 3, note)
        style_cell(c3, fill=fill, font=BODY_FONT,
                   alignment=Alignment(horizontal="left", vertical="top", wrap_text=True), border=THIN_BORDER)
        ws1.row_dimensions[i].height = 80
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 55
    ws1.column_dimensions["C"].width = 25

    # Sheet2: 英文文案
    ws2 = wb.create_sheet("英文文案")
    ws2.append(["英文贴文"])
    ws2.append([data.get("caption", "")])
    ws2.append([])
    ws2.append(["11条Hashtag"])
    hashtags = data.get("hashtags", [])
    if isinstance(hashtags, str):
        hashtags = hashtags.split()
    for tag in hashtags:
        ws2.append([tag])

    # Sheet3: 关键帧描述
    ws3 = wb.create_sheet("关键帧描述")
    ws3.append(["时间节点", "画面描述", "镜头运动", "文字Overlay(英文)", "情绪氛围"])
    for row in data.get("script_rows", []):
        ws3.append(row)

    # Sheet4: 产品匹配
    ws4 = wb.create_sheet("产品匹配")
    ws4.append(["产品名称", "品类", "推荐理由", "产品图"])
    embedded_count = 0
    for p in data.get("matched_products", []):
        ws4.append([p.get("name_en", ""), p.get("category", ""), p.get("match_reason", ""), ""])
        # 插入产品图（如果存在）
        if p.get("image_path") and Path(p["image_path"]).exists():
            img = XLImage(p["image_path"])
            scale_image(img, 150, 150)
            ws4.add_image(img, f"D{ws4.max_row}")
            embedded_count += 1

    logger.info(f"Embedded {embedded_count} product images in content pack")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
    except Exception as e:
        raise RuntimeError(f"Failed to save Excel to {output_path}: {e}")
    return True


def build_design_doc(data: Dict, output_path: str, keyframe_images: List[str]):
    """
    生成设计对接文档.xlsx
    Sheet1: 镜头/画面（含参考示意图）
    Sheet2: 设计参数
    Sheet3: 物料清单
    """
    wb = Workbook()

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    BODY_FONT = Font(name="Arial", size=10)
    thin = Side(style="thin", color="BFBFBF")
    THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_cell(cell, fill=None, font=None, alignment=None, border=None):
        if fill: cell.fill = fill
        if font: cell.font = font
        if alignment: cell.alignment = alignment
        if border: cell.border = border

    # Sheet1: 镜头/画面
    ws1 = wb.active
    ws1.title = "镜头-画面"
    ws1.append(["帧号", "时间", "参考示意图", "设计说明"])
    embedded_count = 0
    for i, (frame_desc, img_path) in enumerate(zip(data.get("script_rows", []), keyframe_images)):
        row_num = i + 2
        time_node = frame_desc[0] if len(frame_desc) > 0 else ""
        design_note = frame_desc[1] if len(frame_desc) > 1 else ""
        ws1.append([f"帧{i+1}", time_node, "", design_note])
        if img_path and Path(img_path).exists():
            img = XLImage(img_path)
            scale_image(img, 200, 355)  # 9:16 竖版
            ws1.add_image(img, f"C{row_num}")
            embedded_count += 1

    logger.info(f"Embedded {embedded_count} keyframe images in design doc")

    # 设置列宽
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 10
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 40

    # Sheet2: 设计参数
    ws2 = wb.create_sheet("设计参数")
    design_params = [
        ("配色方案", data.get("color_palette", "")),
        ("字体建议", data.get("font_suggestion", "")),
        ("排版布局", data.get("layout", "")),
        ("元素位置", data.get("element_positions", "")),
        ("图片规格", "1080x1920 (9:16 竖版)"),
        ("视频规格", "1080x1920, 30fps, 9-30秒"),
    ]
    for row in design_params:
        ws2.append(row)

    # Sheet3: 物料清单
    ws3 = wb.create_sheet("物料清单")
    materials = [
        ("图片物料", "1张 1080x1920 主图"),
        ("视频物料", "1个 1080x1920 视频，9-30秒"),
        ("文案物料", "英文贴文 + 11条Hashtag"),
        ("发布平台", data.get("platform", "Instagram")),
    ]
    for row in materials:
        ws3.append(row)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
    except Exception as e:
        raise RuntimeError(f"Failed to save Excel to {output_path}: {e}")
    return True
